#!/usr/bin/env python3
"""
Analyze Excel structure to map images properly:
- Face photo (individual per participant)
- Skin color reference images (7 types, shared)
- Hair type reference images (10 types, shared)
- Eye color reference images (7 types, shared)
"""

import pandas as pd
import openpyxl
from openpyxl import load_workbook
import json
from pathlib import Path

def analyze_excel_images_detailed(excel_path):
    """Analyze Excel file to understand image placement and types"""

    # Read Excel data
    df = pd.read_excel(excel_path)

    # Load workbook with openpyxl to access images
    wb = load_workbook(excel_path)
    ws = wb.active

    # Image columns mapping
    image_columns = {
        'Which option best matches your skin brightness?(image)': 'skin_brightness_image',
        'Please upload a photo of your facial skin (focus on under-eye area) - Please don\'t take picture with your makeup! ': 'face_photo',
        'What is your Natural-born hair(image)': 'hair_type_image',
        'Which of the following options most matches your natural eye color?(images)': 'eye_color_image'
    }

    # Get column indices
    column_indices = {}
    for col_idx, cell in enumerate(ws[1], 1):  # Header row
        if cell.value:
            for key, img_type in image_columns.items():
                if key in str(cell.value):
                    column_indices[img_type] = col_idx
                    print(f"Found {img_type} at column {col_idx}: {cell.value}")

    # Standard reference images (shared across participants)
    reference_images = {
        'skin_brightness': {
            '1': 'skin_1_fair.png',
            '2': 'skin_2_light.png',
            '3': 'skin_3_light_medium.png',
            '4': 'skin_4_medium.png',
            '5': 'skin_5_medium_deep.png',
            '6': 'skin_6_deep.png',
            '7': 'skin_7_rich.png'
        },
        'hair_types': {
            '1': 'hair_1.png',
            '2a': 'hair_2a.png',
            '2b': 'hair_2b.png',
            '2c': 'hair_2c.png',
            '3a': 'hair_3a.png',
            '3b': 'hair_3b.png',
            '3c': 'hair_3c.png',
            '4a': 'hair_4a.png',
            '4b': 'hair_4b.png',
            '4c': 'hair_4c.png'
        },
        'eye_colors': {
            'Light blue': 'eye_light_blue.png',
            'Blue/Grey': 'eye_blue_grey.png',
            'Green': 'eye_green.png',
            'Hazel': 'eye_hazel.png',
            'Light brown': 'eye_light_brown.png',
            'Medium brown': 'eye_medium_brown.png',
            'Dark brown(Black)': 'eye_dark_brown.png'
        }
    }

    # Process each participant
    participant_mapping = {}

    for idx, row in df.iterrows():
        aid = row['A-ID']
        pid = row['P-ID']
        name = row['What is your full name? (Please write exactly as shown in your ARC or passport)']

        participant_data = {
            'aid': aid,
            'pid': pid,
            'name': name,
            'row': idx + 2,  # Excel row (1-indexed + header)
            'images': {
                'face_photo': None,
                'skin_brightness_ref': None,
                'hair_type_ref': None,
                'eye_color_ref': None
            },
            'data': {
                'skin_brightness': row['밝기판정'],
                'skin_tone': row['톤'],
                'hair_type': row['What is your Natural-born hair(not styled)?[Please select from the 10 options below]'],
                'eye_color': row['Which of the following options most matches your natural eye color?']
            }
        }

        # Map reference images based on data values
        skin_val = str(row['밝기판정']) if pd.notna(row['밝기판정']) else None
        if skin_val and skin_val in reference_images['skin_brightness']:
            participant_data['images']['skin_brightness_ref'] = reference_images['skin_brightness'][skin_val]

        hair_val = str(row['What is your Natural-born hair(not styled)?[Please select from the 10 options below]']).strip().lower() if pd.notna(row['What is your Natural-born hair(not styled)?[Please select from the 10 options below]']) else None
        if hair_val:
            for key in reference_images['hair_types']:
                if key.lower() in hair_val or hair_val in key.lower():
                    participant_data['images']['hair_type_ref'] = reference_images['hair_types'][key]
                    break

        eye_val = str(row['Which of the following options most matches your natural eye color?']).strip() if pd.notna(row['Which of the following options most matches your natural eye color?']) else None
        if eye_val:
            for key in reference_images['eye_colors']:
                if key.lower() in eye_val.lower() or eye_val.lower() in key.lower():
                    participant_data['images']['eye_color_ref'] = reference_images['eye_colors'][key]
                    break

        # Individual face photo - map to actual extracted image
        # Assuming face photos are in sequence with participants
        participant_data['images']['face_photo'] = f"face_photo_{aid}.png"

        participant_mapping[aid] = participant_data

    # Try to match with extracted images
    image_files = list(Path('excel_images').glob('*'))
    print(f"\nTotal extracted images: {len(image_files)}")

    # Group images by type (estimated)
    # Assuming: ~132 face photos + 7 skin + 10 hair + 7 eye = ~156 images
    # The extra images might be duplicates or additional references

    # Simple mapping: assign face photos sequentially
    sorted_aids = sorted(participant_mapping.keys())
    for i, aid in enumerate(sorted_aids):
        if i < len(image_files):
            # Assign face photo from extracted images
            participant_mapping[aid]['images']['face_photo'] = image_files[i].name

    return participant_mapping, reference_images

def save_enhanced_mapping(participant_mapping, reference_images):
    """Save the enhanced mapping with proper image categorization"""

    output = {
        'participants': participant_mapping,
        'reference_images': reference_images,
        'total_participants': len(participant_mapping)
    }

    with open('participant_image_mapping_enhanced.json', 'w', encoding='utf-8') as f:
        json.dump(output, f, ensure_ascii=False, indent=2)

    print(f"\nSaved enhanced mapping for {len(participant_mapping)} participants")

    # Show sample
    sample_aid = list(participant_mapping.keys())[0]
    sample = participant_mapping[sample_aid]
    print(f"\nSample participant {sample_aid}:")
    print(f"  Name: {sample['name']}")
    print(f"  Skin: {sample['data']['skin_brightness']} -> {sample['images']['skin_brightness_ref']}")
    print(f"  Hair: {sample['data']['hair_type']} -> {sample['images']['hair_type_ref']}")
    print(f"  Eye: {sample['data']['eye_color']} -> {sample['images']['eye_color_ref']}")
    print(f"  Face: {sample['images']['face_photo']}")

if __name__ == "__main__":
    excel_path = "makeuptest_AP_Bueatylink_20250927.xlsx"

    print("=== Analyzing Excel Image Structure ===")
    participant_mapping, reference_images = analyze_excel_images_detailed(excel_path)
    save_enhanced_mapping(participant_mapping, reference_images)

    print("\nReference images identified:")
    print(f"  Skin brightness: {len(reference_images['skin_brightness'])} types")
    print(f"  Hair types: {len(reference_images['hair_types'])} types")
    print(f"  Eye colors: {len(reference_images['eye_colors'])} types")